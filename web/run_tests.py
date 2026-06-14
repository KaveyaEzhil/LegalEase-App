import os
import sys
import subprocess
import time
import socket
import openpyxl
import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from test_suite import run_e2e_tests

def wait_for_port(port, host='localhost', timeout=10.0):
    """Wait until a port starts accepting TCP connections."""
    start_time = time.time()
    while True:
        try:
            with socket.create_connection((host, port), timeout=1.0):
                return True
        except socket.error:
            if time.time() - start_time > timeout:
                return False
            time.sleep(0.5)

def shutdown_server(process):
    """Shutdown the Flask server process gracefully."""
    print("Terminating background Flask server process...")
    process.terminate()
    try:
        process.wait(timeout=5)
    except subprocess.TimeoutExpired:
        process.kill()
    print("Flask server shutdown complete.")

def compile_excel_report(test_results, output_path):
    """Compiles the E2E test results into a highly professional Excel workbook."""
    print(f"Creating Excel report at: {output_path}")
    
    wb = openpyxl.Workbook()
    # Rename default sheet
    ws = wb.active
    ws.title = "E2E Test Summary"
    
    # Enable gridlines visibly in Excel
    ws.views.sheetView[0].showGridLines = True
    
    # 1. STYLE DEFINITIONS
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F2937")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_summary_label = Font(name="Segoe UI", size=10, bold=True, color="4B5563")
    font_summary_value = Font(name="Segoe UI", size=10, bold=False, color="111827")
    font_data = Font(name="Segoe UI", size=9, color="1F2937")
    font_data_bold = Font(name="Segoe UI", size=9, bold=True, color="1F2937")
    
    # Fills
    fill_header = PatternFill(start_color="059669", end_color="059669", fill_type="solid") # LegalEase Brand Green
    fill_pass = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Soft Light Green
    fill_fail = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid") # Soft Light Red
    fill_skip = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Soft Light Gray
    
    # Font Colors for Status
    font_pass = Font(name="Segoe UI", size=9, bold=True, color="15803D") # Dark Green
    font_fail = Font(name="Segoe UI", size=9, bold=True, color="B91C1C") # Dark Red
    font_skip = Font(name="Segoe UI", size=9, bold=True, color="4B5563") # Dark Gray
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Borders
    thin_border_side = Side(border_style="thin", color="E5E7EB")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # 2. WRITE SUMMARY HEADER
    ws.merge_cells("A1:G1")
    ws["A1"] = "LegalEase — End-to-End Functionality Testing Report"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    # Metrics calculations
    total_tests = len(test_results)
    passed_tests = sum(1 for r in test_results if r.status == "PASS")
    failed_tests = sum(1 for r in test_results if r.status == "FAIL")
    skipped_tests = sum(1 for r in test_results if r.status == "SKIPPED")
    pass_rate = f"{(passed_tests / total_tests * 100):.1f}%" if total_tests > 0 else "0.0%"
    
    summary_data = [
        ("Test Scope:", "Selenium E2E Functionality Verification Suite", "Total Test Cases:", total_tests),
        ("Execution Time:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Passed Cases:", passed_tests),
        ("Environment:", "Headless Chrome / Local Flask Engine", "Failed Cases:", failed_tests),
        ("Framework Target:", "Flask PWA / Regional Translation", "Pass Rate Metric:", pass_rate)
    ]
    
    for row_idx, (label1, val1, label2, val2) in enumerate(summary_data, start=3):
        ws.cell(row=row_idx, column=1, value=label1).font = font_summary_label
        ws.cell(row=row_idx, column=2, value=val1).font = font_summary_value
        ws.cell(row=row_idx, column=4, value=label2).font = font_summary_label
        ws.cell(row=row_idx, column=5, value=val2).font = font_summary_value
        ws.row_dimensions[row_idx].height = 18
        
    # Add thin spacing row
    ws.row_dimensions[7].height = 10
    
    # 3. WRITE TABLE HEADERS (Row 8)
    headers = [
        "Test ID", "Category", "Test Name", "Description", 
        "Expected Result", "Actual Result", "Status"
    ]
    
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
    ws.row_dimensions[8].height = 24
    
    # 4. WRITE TEST CASE DATA (Row 9 onwards)
    current_row = 9
    for r in test_results:
        # Values
        c_id = ws.cell(row=current_row, column=1, value=r.test_id)
        c_cat = ws.cell(row=current_row, column=2, value=r.category)
        c_name = ws.cell(row=current_row, column=3, value=r.name)
        c_desc = ws.cell(row=current_row, column=4, value=r.description)
        c_exp = ws.cell(row=current_row, column=5, value=r.expected)
        c_act = ws.cell(row=current_row, column=6, value=r.actual)
        c_status = ws.cell(row=current_row, column=7, value=r.status)
        
        # Base formatting
        for cell in [c_id, c_cat, c_name, c_desc, c_exp, c_act, c_status]:
            cell.font = font_data
            cell.border = border_cell
            cell.alignment = align_left_wrap
            
        c_id.alignment = align_center
        c_id.font = font_data_bold
        c_status.alignment = align_center
        
        # Color status column based on outcome
        if r.status == "PASS":
            c_status.fill = fill_pass
            c_status.font = font_pass
        elif r.status == "FAIL":
            c_status.fill = fill_fail
            c_status.font = font_fail
        else:
            c_status.fill = fill_skip
            c_status.font = font_skip
            
        # Adjust row height based on content density
        max_len = max(len(str(r.description)), len(str(r.expected)), len(str(r.actual)))
        ws.row_dimensions[current_row].height = max(18, min(40, max_len // 4))
        
        current_row += 1
        
    # 5. DYNAMIC COLUMN WIDTH AUTO-FIT
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip merged cells and long descriptions/expected/actual for layout spacing
            if cell.row < 8:
                continue
            if cell.column in [4, 5, 6]:  # Description, Expected, Actual columns limit size
                max_len = max(max_len, 25)
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(output_path)
    print("Excel report generated successfully.")

def main():
    print("=" * 60)
    print("          LegalEase E2E Test Execution & Report           ")
    print("=" * 60)
    
    # 1. Start the Flask server in background
    print("Starting Flask application server on localhost:8080...")
    flask_process = subprocess.Popen(
        [sys.executable, 'index.py'],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        bufsize=1
    )
    
    # 2. Wait for server port to open
    print("Waiting for local server port 8080 to become available...")
    if not wait_for_port(8080, timeout=12.0):
        print("Error: Flask server failed to start or bind to port 8080 in time.")
        # Log stdout/stderr of server process
        stdout, stderr = flask_process.communicate()
        print(f"Flask Server stdout:\n{stdout}")
        print(f"Flask Server stderr:\n{stderr}")
        sys.exit(1)
    print("Flask server successfully bound to port 8080.")
    
    # 3. Run E2E tests
    test_results = []
    try:
        test_results = run_e2e_tests('http://localhost:8080')
    except Exception as run_error:
        print(f"Error occurred during test execution: {run_error}")
    finally:
        # 4. Terminate background server process
        shutdown_server(flask_process)
        
    # 5. Compile the results into Excel
    if test_results:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        report_filename = f"E2E_Test_Report_LegalEase_{timestamp}.xlsx"
        compile_excel_report(test_results, report_filename)
        
        # Calculate summary statistics
        total = len(test_results)
        passed = sum(1 for r in test_results if r.status == "PASS")
        failed = sum(1 for r in test_results if r.status == "FAIL")
        skipped = sum(1 for r in test_results if r.status == "SKIPPED")
        
        print("\n" + "=" * 60)
        print("                   TEST EXECUTION SUMMARY                 ")
        print("=" * 60)
        print(f"Total Test Cases Executed : {total}")
        print(f"Passed Cases              : {passed} (Green)")
        print(f"Failed Cases              : {failed} (Red)")
        print(f"Skipped Cases             : {skipped} (Gray)")
        print(f"Pass Rate Percentage      : {(passed/total*100):.1f}%")
        print(f"Report File Created       : {os.path.abspath(report_filename)}")
        print("=" * 60)
    else:
        print("No test results gathered. Excel report was not generated.")

if __name__ == "__main__":
    main()
